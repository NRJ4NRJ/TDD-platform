import json
from pathlib import Path
from openpyxl import load_workbook

path = Path('20260319_TDD_Projet Mimosa_Vinterne (1).xlsx')
wb = load_workbook(path, data_only=True)

report = {
    'source_file': path.name,
    'sheet_names': wb.sheetnames,
    'project_metadata': {},
    'sections': [],
}

label_map = {
    'Nom du projet': 'name',
    'SPV': 'spv',
    'Commune': 'commune',
    'Département': 'department',
    'Région': 'region',
    'Pays': 'country',
    'Statut du projet': 'project_status',
    'Equipement': 'equipment',
    'Puissance totale ': 'capacity',
}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = [tuple(cell if cell != '\xa0' else None for cell in row) for row in ws.iter_rows(values_only=True)]
    non_empty = [row for row in rows if any(cell is not None for cell in row)]
    if sheet_name == 'Executive summary':
        for row in non_empty:
            if len(row) >= 5 and row[1] and row[4] and row[0] is None:
                label = str(row[1]).strip()
                value = row[4]
                if label in label_map:
                    report['project_metadata'][label_map[label]] = str(value).strip() if not isinstance(value, (int, float)) else value
                else:
                    report['project_metadata'][label] = str(value).strip() if not isinstance(value, (int, float)) else value
    elif sheet_name in ['1. Permitting', '2. Design', '3. Raccordement', '4. AO CRE', '5. FM ', '6. EYAr', '7. Additional docs']:
        header = None
        header_index = None
        for idx, row in enumerate(non_empty):
            if row and any(isinstance(c, str) and c.strip() == 'Ref' for c in row):
                header = [str(c).strip() if c is not None else None for c in row]
                header_index = idx
                break
        if header is None:
            continue
        section = {
            'sheet_name': sheet_name,
            'rows': [],
        }
        for row in non_empty[header_index + 1:]:
            if not any(cell is not None for cell in row):
                continue
            data = {}
            for col_idx, head in enumerate(header):
                if head is None:
                    continue
                val = row[col_idx] if col_idx < len(row) else None
                if isinstance(val, str):
                    val = val.strip()
                    if val == '\xa0':
                        val = None
                data[head] = val
            if data.get('Subject') or data.get('Sub-subject') or data.get('Description'):
                section['rows'].append(data)
        report['sections'].append(section)

out = Path('src/data/mimosa-report.json')
out.parent.mkdir(parents=True, exist_ok=True)
with out.open('w', encoding='utf-8') as f:
    json.dump(report, f, indent=2, ensure_ascii=False)
print('wrote', out)
